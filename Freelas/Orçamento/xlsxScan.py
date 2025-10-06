import os
import shutil
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter

def criar_copia_editavel(arquivo_origem, arquivo_destino=None):
    """
    Cria uma cópia editável de um arquivo Excel mantendo toda a formatação original
    
    Args:
        arquivo_origem (str): Caminho do arquivo Excel original
        arquivo_destino (str): Caminho do arquivo de destino (opcional)
    
    Returns:
        str: Caminho do arquivo de destino criado
    """
    
    if arquivo_destino is None:
        nome, extensao = os.path.splitext(arquivo_origem)
        arquivo_destino = f"{nome}_copia_editavel{extensao}"
    
    # Se o destino já existe, remover
    if os.path.exists(arquivo_destino):
        os.remove(arquivo_destino)
    
    # Criar cópia exata do arquivo original
    shutil.copy2(arquivo_origem, arquivo_destino)
    print(f"Cópia criada: {arquivo_destino}")
    
    return arquivo_destino

def carregar_dados_com_formato(arquivo, aba=None):
    """
    Carrega dados do Excel mantendo referência para formatação
    
    Returns:
        tuple: (dataframe, workbook, worksheet) para edição posterior
    """
    wb = load_workbook(arquivo)
    
    if aba is None:
        ws = wb.active
    else:
        ws = wb[aba]
    
    # Converter para DataFrame
    dados = []
    for row in ws.iter_rows(values_only=True):
        dados.append(row)
    
    # Criar DataFrame (assume primeira linha como cabeçalho)
    df = pd.DataFrame(dados[1:], columns=dados[0]) if dados else pd.DataFrame()
    
    return df, wb, ws

def salvar_dados_com_formato(df, wb, ws, arquivo_destino):
    """
    Salva dados mantendo a formatação original
    """
    # Limpar dados existentes (mantendo formatação)
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            try:
                celula = encontrar_celula_editavel(ws, row, col)
                if celula.row == row and celula.column == col:  # Só limpa a célula âncora
                    celula.value = None
            except:
                continue
    
    # Escrever novos dados
    # Cabeçalhos
    for col_idx, col_name in enumerate(df.columns, 1):
        celula = encontrar_celula_editavel(ws, 1, col_idx)
        celula.value = col_name

    # Dados
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            celula = encontrar_celula_editavel(ws, row_idx, col_idx)  # Corrigido: row_idx em vez de 1
            celula.value = value  # Corrigido: celula.value em vez de celular.value

    # Salvar
    wb.save(arquivo_destino)
    print(f"Dados salvos com formatação: {arquivo_destino}")

def adicionar_linhas_simples(arquivo, aba=None, quantidade=1, posicao="final"):
    """
    Adiciona linhas vazias mantendo a formatação
    """
    wb = load_workbook(arquivo)
    ws = wb[aba] if aba else wb.active

    if posicao == "final":
        linha_alvo = ws.max_row + 1
    else:
        linha_alvo = posicao if isinstance(posicao, int) else 2

    # Adicionar múltiplas linhas se necessário
    for _ in range(quantidade):
        ws.insert_rows(linha_alvo)
    
    wb.save(arquivo)
    print(f"Adicionadas {quantidade} linha(s) na posição {linha_alvo}")

def copiar_formatacao_linha(arquivo, aba=None, linha_base=None, linha_destino=None):
    """
    Copia formatação de uma linha base para uma linha destino, incluindo mesclagem e bordas
    """

# Função mantida para compatibilidade (usando o novo nome internamente)
def is_merged(ws,row,col):
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return True
    return False

def encontrar_mesclagem(ws):
    for merged_range in ws.merged_cells.ranges:
        print(f"Mesclado: {merged_range}")

def encontrar_celula_editavel(ws, row, col):
    """
    Encontra a célula editável em uma posição, lidando com células mescladas
    """
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return ws.cell(merged_range.min_row, merged_range.min_col)
    return ws.cell(row, col)

def remover_linhas(arquivo, aba=None, linhas=None):
    """
    Remove linhas específicas
    """
    wb = load_workbook(arquivo)
    ws = wb[aba] if aba else wb.active
    
    if isinstance(linhas, int):
        linhas = [linhas]
    
    # Ordenar e remover da maior para a menor linha
    for linha in sorted(linhas, reverse=True):
        ws.delete_rows(linha)
    
    wb.save(arquivo)
    print(f"Linhas removidas: {linhas}")

def editar_celula(arquivo, aba=None, linha=1, coluna=1, valor=None):
    """
    Edita uma célula específica mantendo a formatação
    """
    wb = load_workbook(arquivo)
    ws = wb[aba] if aba else wb.active
    
    celula = encontrar_celula_editavel(ws, linha, coluna)
    celula.value = valor
    
    wb.save(arquivo)

# EXEMPLO PRÁTICO DE USO
def exemplo_completo():
    """
    Exemplo completo de como usar as funções
    """
    
    arquivo_editavel = criar_copia_editavel("ORÇAMENTO DE EVENTOS.xlsx")
    
    # 3. CARREGAR DADOS PARA EDIÇÃO
    df, wb, ws = carregar_dados_com_formato(arquivo_editavel, "ENVIO P CLIENTE")
    print("\nDados originais:")
    print(df)
    
    editar_celula(arquivo_editavel, "ENVIO P CLIENTE", linha=14, coluna=8, valor="R$ 2000")  # Alterar quantidade do Mouse
    
    # 5. VERIFICAR RESULTADO FINAL
    df_final, _, _ = carregar_dados_com_formato(arquivo_editavel, "ENVIO P CLIENTE")
    print("\nDados após edição:")
    print(df_final)
    
    print(f"\n✅ Arquivo editável pronto: {arquivo_editavel}")
    print("Toda a formatação foi mantida!")

# USO RÁPIDO
def uso_rapido(arquivo_original):
    """
    Uso rápido para seu arquivo específico
    """
    # Criar cópia editável
    arquivo_copia = criar_copia_editavel(arquivo_original)
    
    print(f"\n🎯 Agora você pode editar programaticamente:")
    print(f"Arquivo original: {arquivo_original}")
    print(f"Arquivo editável: {arquivo_copia}")
    print(f"\n📝 Funções disponíveis:")
    print(f"- adicionar_linhas('{arquivo_copia}', dados=[...])")
    print(f"- remover_linhas('{arquivo_copia}', linhas=[2, 3, 4])")
    print(f"- editar_celula('{arquivo_copia}', linha=1, coluna=1, valor='Novo Valor')")
    
    return arquivo_copia

if __name__ == "__main__":
    # Opção 1: Executar exemplo completo
    exemplo_completo()
    
    # Opção 2: Uso rápido com seu arquivo (descomente a linha abaixo)
    #uso_rapido("ORÇAMENTO DE EVENTOS.xlsx")
