from openpyxl.styles import Border, Font, Alignment, Side
from openpyxl import Workbook
from openpyxl.drawing.image import Image

def add_border(cell, top=None, bottom=None, left=None, right=None):
    borda_atual = cell.border
    nova_borda = Border(
        top=top if top else borda_atual.top,
        bottom=bottom if bottom else borda_atual.bottom,
        left=left if left else borda_atual.left,
        right=right if right else borda_atual.right,
    )
    cell.border = nova_borda

def gera_arquivo(nome_arquivo: str, nome_aba: str, cliente: str, obra: str, local: str, nega: str, area: str, be: str, estaca: str, secao: str, data_inicial: str, data_final: str, comp_cravado: float, peso_martelo: str, altura_queda: float, golpes: list):
    wb = Workbook()
    wb.remove(wb.active)
    gera_aba(wb, nome_aba, cliente, obra, local, nega, area, be, estaca, secao, data_inicial, data_final, comp_cravado, peso_martelo, altura_queda, golpes)
    wb.save(nome_arquivo)
    return wb

def gera_aba(wb, nome_aba: str, cliente: str, obra: str, local: str, nega: str, area: str, be: str, estaca: str, secao: str, data_inicial: str, data_final: str, comp_cravado: float, peso_martelo: str, altura_queda: float, golpes: list):
    borda_fina = Side(style="thin")
    borda_grossa = Side(style="medium")
    ws = wb.create_sheet(nome_aba)
    ws.merge_cells('A1:L3')

    ws['A4'] = f'Cliente: {cliente}'
    ws['A5'] = f'Obra: {obra}'
    ws['A6'] = f'Local: {local}'
    ws['A7'] = 'Comp. Cravado (m):'   ;ws['D7'] = comp_cravado
    ws['A8'] = 'Peso do Martelo (kg):';ws['D8'] = peso_martelo
    ws['A9'] = 'Altura de Queda (m):' ;ws['D9'] = altura_queda
    for row in ws['D7:D9']:
        for cell in row:
            cell.font = Font(name='Arial',
                             size=11)
            cell.alignment = Alignment(horizontal='right')
    ws['A10'] = f'Nega (mm): {nega}'
    for row in ws['A10:L10']:
        for cell in row:
            add_border(cell, bottom=borda_fina)

    ws['G5'] = 'Área:'        ;ws['H5'] = area
    ws['G6'] = 'BE:'          ;ws['H6'] = be
    ws['G7'] = 'Estaca:'      ;ws['H7'] = estaca
    ws['G8'] = 'Seção:'       ;ws['H8'] = secao
    ws['G9'] = 'Data Inicial:';ws['H9'] = data_inicial
    ws['G10'] = 'Data Final'  ;ws['H10'] = data_final
    for row in ws['G5:G10']:
        for cell in row:
            cell.font = Font(name="Arial",
                             bold=True,
                             size=11)
            cell.alignment = Alignment(horizontal='right')
    for row in ws['H5:H10']:
        for cell in row:
            cell.font = Font(name='Arial',
                             size=11)
            cell.alignment = Alignment(horizontal='left')
    ws['A11'] = 'Profun-'
    ws['A11'].border = Border(top=borda_fina, left=borda_grossa, right=borda_fina)
    ws['A12'] = 'didade'
    ws['A12'].border = Border(left=borda_grossa, right=borda_fina)
    ws['A13'] = '(m)'
    ws['A13'].border = Border(bottom=borda_fina, left=borda_grossa, right=borda_fina)
    ws['B11'] = 'Número'
    ws['B11'].border = Border(top=borda_fina, left=borda_fina, right=borda_fina)
    ws['B12'] = 'de'
    ws['B12'].border = Border(left=borda_fina, right=borda_fina)
    ws['B13'] = 'Golpes'
    ws['B13'].border = Border(bottom=borda_fina, left=borda_fina, right=borda_fina)
    for row in ws['A4:B13']:
        for cell in row:
            cell.font = Font(name="Arial",
                             bold=True,
                             size=11)
    for row in ws['A11:B13']:
        for cell in row:
            cell.alignment = Alignment(horizontal='center')
    for i in range(4,13+1):
        ws.row_dimensions[i].height = 15
    ws.row_dimensions[3].height = 5
    ws.column_dimensions['J'].width = 10
    ws.column_dimensions['C'].width = 3
    ws.column_dimensions['I'].width = 1
    ws.column_dimensions['K'].width = 1

    for i in range(int(comp_cravado+1)):
        if i == int(comp_cravado) and comp_cravado != int(comp_cravado):
            ws[f'A{14+1+i}'] = comp_cravado
        ws[f'A{14+i}'] = i
    for i, golpe in enumerate(golpes):
        ws[f'B{14+i}'] = golpe

    for row in ws['A14:B70']:
        for cell in row:
            add_border(cell,
                       top=borda_fina,
                       bottom=borda_fina,
                       right=borda_fina)
    for row in ws['A14:B70']:
        for cell in row:
            cell.font = Font(size=8,
                             name='Arial')
            cell.alignment = Alignment(horizontal='center', vertical='center')
    for i in range(14,71+1):
        ws.row_dimensions[i].height = 8.5

    for row in ws['A72:L72']:
        for cell in row:
            add_border(cell, top=borda_fina)

    ws.merge_cells('A75:E75')
    for row in ws['A75:E75']:
        for cell in row:
            add_border(cell, top=borda_fina)

    ws.merge_cells('G75:J75')
    for row in ws['G75:J75']:
        for cell in row:
            add_border(cell, top=borda_fina)
    ws['A75'] = 'FN Sondagens, fundações e obras especiais'
    ws['G75'] = 'Contratante'
    for row in ws['A75:J75']:
        for cell in row:
            cell.font = Font(name='Arial',
                             size=10)
            cell.alignment = Alignment(horizontal='center')
    ws.merge_cells('A77:L77')
    ws['A77'] = 'Av. Gov. Jose Malcher, Nº 168, sala 110 - Nazaré - Belém / PA - CEP: 66.035-065'
    ws.merge_cells('A78:L78')
    ws['A78'] = 'Fone: (91) 3039-7200 / 99349-2626 / 99984-0640'
    ws.merge_cells('A79:L79')
    ws['A79'] = ' Site: www.fnengenharia.eng.br / email: contato@fnengenharia.eng.br / fncresponeto@gmail.com'
    for row in ws['A77:L79']:
        for cell in row:
            cell.font = Font(name='Arial',
                             size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[72].height = 5
    ws.row_dimensions[73].height = 15
    ws.row_dimensions[74].height = 15
    for i in range(75,79+1):
        ws.row_dimensions[i].height = 12
 
    fn_logo = Image("fn_logo.png")
    fn_logo.width = 175
    fn_logo.height = 40
    ws.add_image(fn_logo, "A1")

    assinatura = Image('assinatura.png')
    assinatura.width = 175
    assinatura.height = 35
    ws.add_image(assinatura, 'B73')

    insta = Image('insta_fn.jpg')
    insta.width = 125
    insta.height = 12
    ws.add_image(insta, 'J78')

    for row in ws['A1:L1']:
        for cell in row:
            add_border(cell, top=borda_grossa)
    for row in ws['A79:L79']:
        for cell in row:
            add_border(cell, bottom=borda_grossa)
    for row in ws['L1:L79']:
        for cell in row:
            add_border(cell, right=borda_grossa)
    for row in ws['A1:A79']:
        for cell in row:
            add_border(cell, left=borda_grossa)
    for row in ws['A3:L3']:
        for cell in row:
            add_border(cell, bottom=borda_fina)

    ws.sheet_view.showGridLines = False
    ws.print_area = "A1:L79"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    '''ws.row_dimensions.group(80,1048576, hidden=True)
    ws.column_dimensions.group('M','XFD', hidden=True)'''

def main():
    golpes = []
    for i in range(32+1):
        golpes.append(i)
    gera_arquivo(nome_arquivo='teste.xlsx',
                 nome_aba='aba',
                 cliente='testison',
                 obra='obra teste',
                 local='minha casa',
                 nega='25/04/2002',
                 area='seila',
                 be='outro que nao sei',
                 estaca='estaca teste',
                 secao='secao teste',
                 data_inicial='25/05/2020',
                 data_final='25/05/2021',
                 comp_cravado=32.0,
                 peso_martelo="4,500",
                 altura_queda=1.00,
                 golpes=golpes)
if __name__ == '__main__': main();
