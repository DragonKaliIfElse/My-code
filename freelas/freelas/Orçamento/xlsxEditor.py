from openpyxl import load_workbook
from copy import copy # Importante para garantir a cópia de objetos complexos
import re
import pathlib
import os
import shutil
from openpyxl.utils import range_boundaries

def unmerge_line(ws,line):
    merged = list(ws.merged_cells.ranges)
    for merge in merged:
        linha = re.split(r'[A-Z]+', str(merge))[1]
        linha = linha.split(':')[0]
        if str(line) == str(linha):
            ws.unmerge_cells(str(merge))
            print(merge)

def unmerge_line_horizontal(ws, line):
    merged_ranges_to_unmerge = list(ws.merged_cells.ranges)
    target_line = str(line)
    
    print(f"Buscando desmesclar intervalos horizontais na linha: {target_line}")
    print("-" * 40)
    
    for merge_range_str in merged_ranges_to_unmerge:
        
        try:
            # 3. Extrai o número da linha inicial
            # re.split(r'[A-Z]+', 'A9:D9') -> ['', '9:', '9']
            # [1] -> '9:'
            linha_extraida = re.split(r'[A-Z]+', str(merge_range_str))[1]
            
            # '9:'.split(':')[0] -> '9'
            linha_extraida = linha_extraida.split(':')[0]

            # 4. Compara com a linha alvo
            if target_line == linha_extraida:
                
                # 5. Opcional, mas recomendado: Uma verificação rápida para garantir que é horizontal
                #    Esta linha verifica se o intervalo começa e termina na mesma linha.
                min_col, min_row, max_col, max_row = range_boundaries(str(merge_range_str))
                if min_row == max_row:
                    
                    ws.unmerge_cells(str(merge_range_str))
                    print(f"Desmesclado: {merge_range_str}")
                else:
                    # Se você tem certeza que é apenas horizontal, essa parte não será executada.
                    print(f"Atenção: O intervalo {merge_range_str} não é puramente horizontal e foi ignorado.")

        except IndexError:
            print(f"Erro ao processar o intervalo {merge_range_str}. Não segue o padrão esperado.")

def cop_sheet(arquivo):
    arq = arquivo.split('.')[0]
    nome_copia = f'PROPOSTA.xlsx'
    try:
        shutil.copy2(arquivo, nome_copia)
        return nome_copia
    except FileNotFoundError:
        print(f"Erro: O arquivo original '{arquivo}' não foi encontrado.")
        return None

def copy_cell_format(ws,row1,column1,row2,column2,cp_value=False):
    celula_origem = ws.cell(row=row1,column=column1)
    celula_destino = ws.cell(row=row2,column=column2)

    # Copiar valor
    if cp_value is True:
        
        celula_destino.value = celula_origem.value
    
    if celula_origem.has_style:
        celula_destino.font = copy(celula_origem.font)
        celula_destino.border = copy(celula_origem.border)
        celula_destino.fill = copy(celula_origem.fill)
        celula_destino.number_format = copy(celula_origem.number_format)
        celula_destino.protection = copy(celula_origem.protection)
        celula_destino.alignment = copy(celula_origem.alignment)

def copy_line_format(ws,row1,row2,column_max,cp_value=False):
    colunas1_mescladas = []
    merged_ranges_to_unmerge = list(ws.merged_cells.ranges)
    unmerge_line_horizontal(ws,row2)
    for merge_range_str in merged_ranges_to_unmerge:
        try:
            linha_extraida = re.split(r'[A-Z]+', str(merge_range_str))[1]
            linha_extraida = linha_extraida.split(':')[0]
            if str(row1) == str(linha_extraida):
                colunas_destino = change_line(str(merge_range_str),str(row2))            
                print(f'essa é a {colunas_destino}')
                ws.merge_cells(str(colunas_destino))
        except Exception as e:
            print('erro no merge_cell',e)
            pass
    for column in range(1,column_max):
        celula_origem = encontrar_celula_editavel(ws,row1,column)
        celula_destino = encontrar_celula_editavel(ws,row2,column)

        # Copiar valor
        if cp_value is True:
            celula_destino.value = celula_origem.value
        
        if celula_origem.has_style:
            celula_destino.font = copy(celula_origem.font)
            celula_destino.border = copy(celula_origem.border)
            celula_destino.fill = copy(celula_origem.fill)
            celula_destino.number_format = copy(celula_origem.number_format)
            celula_destino.protection = copy(celula_origem.protection)
            celula_destino.alignment = copy(celula_origem.alignment)

def store_line_data(ws, row1, column_max):
    line_data = {
        'cell_data': [],
        'merged_cells': []
    }
    for column in range(1, column_max + 1): # Incluindo column_max
        celula_origem = encontrar_celula_editavel(ws, row1, column)
        
        cell_info = {
            'column': column,
            'value': celula_origem.value,
            'has_style': celula_origem.has_style,
        }
        if celula_origem.has_style:
            cell_info['font'] = copy(celula_origem.font)
            cell_info['border'] = copy(celula_origem.border)
            cell_info['fill'] = copy(celula_origem.fill)
            cell_info['number_format'] = copy(celula_origem.number_format)
            cell_info['protection'] = copy(celula_origem.protection)
            cell_info['alignment'] = copy(celula_origem.alignment)
        line_data['cell_data'].append(cell_info)

    for merge_range in list(ws.merged_cells.ranges):
        if merge_range.min_row == row1:
            line_data['merged_cells'].append(str(merge_range))
            
    return line_data

def apply_line_data(ws, row2, stored_data, cp_value=False):
    unmerge_line_horizontal(ws,row2)
    if len(stored_data['merged_cells']) > 0:
        for merge_range_str in stored_data['merged_cells']:
            try:
                colunas_destino = change_line(merge_range_str, str(row2))
                print(f'Mesclando na linha {row2}: {colunas_destino}')
                ws.merge_cells(str(colunas_destino))
            except Exception as e:
                print('Erro ao aplicar merge_cell', e)
    
    for cell_info in stored_data['cell_data']:
        column = cell_info['column']
        celula_destino = encontrar_celula_editavel(ws, row2, column)
        
        if cp_value and cell_info['value'] is not None:
            celula_destino.value = cell_info['value']
        
        if cell_info['has_style']:
            celula_destino.font = cell_info['font']
            celula_destino.border = cell_info['border']
            celula_destino.fill = cell_info['fill']
            celula_destino.number_format = cell_info['number_format']
            celula_destino.protection = cell_info['protection']
            celula_destino.alignment = cell_info['alignment']

def encontrar_celula_editavel(ws, row, col):
    """
    Encontra a célula editável em uma posição, lidando com células mescladas
    """
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return ws.cell(merged_range.min_row, merged_range.min_col)
    return ws.cell(row, col)

def is_merged(ws,row,col):
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row and
            merged_range.min_col <= col <= merged_range.max_col):
            return True
    return False

def change_line(cell_range_str, new_line):
    new_line_str = str(new_line)
    regex = r"([A-Z])\d+" 
    substituicao = r"\g<1>" + new_line_str
    novo_range = re.sub(regex, substituicao, cell_range_str)
    return novo_range

# Teste com um número de dois dígitos que causaria o erro:
def main():
    arquivo = 'ORÇAMENTO DE EVENTOS.xlsx'
    aba  = 'ENVIO P CLIENTE'
    editavel = cop_sheet(arquivo)

    wb = load_workbook(editavel)
    ws = wb[aba]
    line_data = store_line_data(ws,49,10)
    apply_line_data(ws,98,line_data,True)
    wb.save(editavel)
if __name__ == '__main__':main();
