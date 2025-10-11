import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Dict, List, Tuple, Optional, Any
import pandas as pd
import re
from pathlib import Path
import xlsxEditor as xe
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from datetime import datetime


class DatabaseViewer:
    """Classe para visualização e edição da base de dados"""
    
    def __init__(self, parent, arquivo):
        self.parent = parent
        self.arquivo = arquivo
        self.df = None
        self.df_original = None
        self.tree = None
        self.frame = None
        
        self.criar_interface()
        self.carregar_dados()
    
    def criar_interface(self):
        """Cria a interface do visualizador de base de dados"""
        self.frame = ttk.Frame(self.parent)
        
        # Frame de controles
        frame_controles = ttk.Frame(self.frame)
        frame_controles.pack(fill="x", padx=10, pady=5)
        
        ttk.Button(frame_controles, text="Recarregar Dados", 
                  command=self.carregar_dados).pack(side="left", padx=5)
        ttk.Button(frame_controles, text="Salvar Alterações", 
                  command=self.salvar_alteracoes).pack(side="left", padx=5)
        ttk.Button(frame_controles, text="Adicionar Linha", 
                  command=self.adicionar_linha).pack(side="left", padx=5)
        ttk.Button(frame_controles, text="Editar Linha Selecionada", 
                  command=self.editar_linha).pack(side="left", padx=5)
        ttk.Button(frame_controles, text="Remover Linha Selecionada", 
                  command=self.remover_linha).pack(side="left", padx=5)
        ttk.Button(frame_controles, text="Gerar Relatório", 
                  command=self.gerar_relatorio).pack(side="left", padx=5)
        
        # Frame de filtros
        frame_filtros = ttk.Frame(self.frame)
        frame_filtros.pack(fill="x", padx=10, pady=5)
        
        ttk.Label(frame_filtros, text="Filtrar por Evento:").pack(side="left", padx=5)
        self.filtro_evento = ttk.Entry(frame_filtros, width=20)
        self.filtro_evento.pack(side="left", padx=5)
        
        ttk.Label(frame_filtros, text="Data Início:").pack(side="left", padx=5)
        self.filtro_data_inicio = ttk.Entry(frame_filtros, width=15)
        self.filtro_data_inicio.pack(side="left", padx=5)
        
        ttk.Button(frame_filtros, text="Aplicar Filtro", 
                  command=self.aplicar_filtro).pack(side="left", padx=5)
        ttk.Button(frame_filtros, text="Limpar Filtro", 
                  command=self.limpar_filtro).pack(side="left", padx=5)
        
        # Frame da tabela
        frame_tabela = ttk.Frame(self.frame)
        frame_tabela.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Criar Treeview com scrollbars
        self.criar_treeview(frame_tabela)
    
    def criar_treeview(self, parent):
        """Cria o widget Treeview para exibir os dados"""
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(parent, orient="vertical")
        h_scrollbar = ttk.Scrollbar(parent, orient="horizontal")
        
        # Treeview
        self.tree = ttk.Treeview(parent, 
                                yscrollcommand=v_scrollbar.set,
                                xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.config(command=self.tree.yview)
        h_scrollbar.config(command=self.tree.xview)
        
        # Layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        
        parent.grid_rowconfigure(0, weight=1)
        parent.grid_columnconfigure(0, weight=1)
        
        # Bind para edição
        self.tree.bind("<Double-1>", self.editar_celula)
    
    def carregar_dados(self):
        """Carrega os dados da planilha"""
        try:
            # Tentar carregar a aba LANÇAMENTOS
            try:
                self.df = pd.read_excel(self.arquivo, sheet_name="LANÇAMENTOS")
                # Verificar e corrigir nomes de colunas
                self.df.columns = self.df.columns.str.strip()
                self.df_original = self.df.copy()
                print("Colunas carregadas:", list(self.df.columns))
            except Exception as e:
                print(f"Erro ao carregar LANÇAMENTOS: {e}")
                # Se não existir, criar DataFrame vazio com colunas padrão
                colunas = [
                    "Inicio", "Termino", "Evento", "Categoria", "Item", 
                    "tipo", "Quantidade", "Diária", "Subtotal", "Total", "Incluir"
                ]
                self.df = pd.DataFrame(columns=colunas)
                self.df_original = self.df.copy()
            
            self.atualizar_treeview()
            messagebox.showinfo("Sucesso", "Dados carregados com sucesso!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar dados: {str(e)}")
    
    def aplicar_filtro(self):
        """Aplica filtros nos dados"""
        try:
            filtro_evento = self.filtro_evento.get().strip()
            filtro_data = self.filtro_data_inicio.get().strip()
            
            if not filtro_evento and not filtro_data:
                self.df = self.df_original.copy()
            else:
                self.df = self.df_original.copy()
                
                if filtro_evento:
                    mask = self.df['Evento'].astype(str).str.contains(filtro_evento, case=False, na=False)
                    self.df = self.df[mask]
                
                if filtro_data:
                    mask = self.df['Inicio'].astype(str).str.contains(filtro_data, na=False)
                    self.df = self.df[mask]
            
            self.atualizar_treeview()
            messagebox.showinfo("Sucesso", f"Filtro aplicado! {len(self.df)} registros encontrados.")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao aplicar filtro: {str(e)}")
    
    def limpar_filtro(self):
        """Limpa todos os filtros"""
        self.filtro_evento.delete(0, tk.END)
        self.filtro_data_inicio.delete(0, tk.END)
        self.df = self.df_original.copy()
        self.atualizar_treeview()
        messagebox.showinfo("Sucesso", "Filtros limpos!")
    
    def atualizar_treeview(self):
        """Atualiza o Treeview com os dados do DataFrame"""
        try:
            # Limpar treeview existente
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            if self.df.empty:
                return
            
            # Configurar colunas
            colunas = list(self.df.columns)
            self.tree["columns"] = colunas
            self.tree["show"] = "headings"
            
            # Configurar cabeçalhos
            for coluna in colunas:
                self.tree.heading(coluna, text=coluna)
                self.tree.column(coluna, width=100, minwidth=50)
            
            # Adicionar dados
            for index, row in self.df.iterrows():
                valores = [str(row[col]) if pd.notna(row[col]) else "" for col in colunas]
                self.tree.insert("", "end", values=valores, iid=str(index))
        except Exception as e:
            print(f"Erro ao atualizar treeview: {e}")
    
    def editar_celula(self, event):
        """Permite editar células individualmente"""
        try:
            selecionados = self.tree.selection()
            if not selecionados:
                return
                
            item = selecionados[0]
            coluna = self.tree.identify_column(event.x)
            if not coluna.startswith('#'):
                return
                
            col_index = int(coluna[1:]) - 1
            
            if col_index >= len(self.df.columns):
                return
                
            col_name = self.df.columns[col_index]
            
            # Obter valor atual
            valores_atuais = self.tree.item(item, "values")
            valor_atual = valores_atuais[col_index]
            
            # Criar janela de edição
            self.criar_janela_edicao(item, col_index, col_name, valor_atual)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao editar célula: {str(e)}")
    
    def editar_linha(self):
        """Edita a linha completa selecionada"""
        try:
            selecionados = self.tree.selection()
            if not selecionados:
                messagebox.showwarning("Aviso", "Selecione uma linha para editar.")
                return
                
            item = selecionados[0]
            index = int(item)
            
            # Verificar se o índice existe no DataFrame
            if index not in self.df.index:
                messagebox.showerror("Erro", "Índice da linha não encontrado no DataFrame.")
                return
            
            # Obter dados atuais da linha
            linha_data = self.df.loc[index]
            
            # Criar janela de edição completa
            self.criar_janela_edicao_linha(item, linha_data)
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao editar linha: {str(e)}")
    
    def criar_janela_edicao(self, item, col_index, col_name, valor_atual):
        """Cria janela para edição de célula"""
        try:
            janela_edicao = tk.Toplevel(self.frame)
            janela_edicao.title(f"Editar {col_name}")
            janela_edicao.geometry("300x150")
            janela_edicao.transient(self.frame)
            
            # Centralizar na tela
            janela_edicao.update_idletasks()
            x = self.frame.winfo_rootx() + (self.frame.winfo_width() // 2) - (300 // 2)
            y = self.frame.winfo_rooty() + (self.frame.winfo_height() // 2) - (150 // 2)
            janela_edicao.geometry(f"+{x}+{y}")
            
            ttk.Label(janela_edicao, text=f"Editar {col_name}:").pack(pady=10)
            
            entry_valor = ttk.Entry(janela_edicao, width=30)
            entry_valor.pack(pady=5)
            entry_valor.insert(0, valor_atual)
            entry_valor.focus()
            
            def confirmar_edicao():
                try:
                    novo_valor = entry_valor.get()
                    
                    # Atualizar Treeview
                    valores = list(self.tree.item(item, "values"))
                    valores[col_index] = novo_valor
                    self.tree.item(item, values=valores)
                    
                    # Atualizar DataFrame
                    index = int(item)
                    self.df.at[index, col_name] = novo_valor
                    
                    janela_edicao.destroy()
                    messagebox.showinfo("Sucesso", "Célula editada com sucesso!")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao confirmar edição: {str(e)}")
            
            frame_botoes = ttk.Frame(janela_edicao)
            frame_botoes.pack(pady=10)
            
            ttk.Button(frame_botoes, text="Confirmar", 
                      command=confirmar_edicao).pack(side="left", padx=5)
            ttk.Button(frame_botoes, text="Cancelar", 
                      command=janela_edicao.destroy).pack(side="left", padx=5)
            
            entry_valor.bind("<Return>", lambda e: confirmar_edicao())
            
            # Usar protocolo para evitar problemas com grab_set
            janela_edicao.protocol("WM_DELETE_WINDOW", janela_edicao.destroy)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar janela de edição: {str(e)}")
    
    def criar_janela_edicao_linha(self, item, linha_data):
        """Cria janela para edição completa da linha"""
        try:
            janela_edicao = tk.Toplevel(self.frame)
            janela_edicao.title("Editar Linha Completa")
            janela_edicao.geometry("600x400")
            janela_edicao.transient(self.frame)
            
            # Centralizar na tela
            janela_edicao.update_idletasks()
            x = self.frame.winfo_rootx() + (self.frame.winfo_width() // 2) - (600 // 2)
            y = self.frame.winfo_rooty() + (self.frame.winfo_height() // 2) - (400 // 2)
            janela_edicao.geometry(f"+{x}+{y}")
            
            # Frame com scroll
            container = ttk.Frame(janela_edicao)
            container.pack(fill="both", expand=True, padx=10, pady=10)
            
            canvas = tk.Canvas(container)
            scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Campos de edição
            entries = {}
            for i, coluna in enumerate(self.df.columns):
                ttk.Label(scrollable_frame, text=f"{coluna}:").grid(row=i, column=0, padx=5, pady=5, sticky="e")
                entry = ttk.Entry(scrollable_frame, width=40)
                entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
                valor = str(linha_data[coluna]) if pd.notna(linha_data[coluna]) else ""
                entry.insert(0, valor)
                entries[coluna] = entry
            
            scrollable_frame.columnconfigure(1, weight=1)
            
            def confirmar_edicao():
                try:
                    # Atualizar DataFrame
                    index = int(item)
                    for coluna, entry in entries.items():
                        self.df.at[index, coluna] = entry.get()
                    
                    # Atualizar Treeview
                    novos_valores = [entry.get() for entry in entries.values()]
                    self.tree.item(item, values=novos_valores)
                    
                    janela_edicao.destroy()
                    messagebox.showinfo("Sucesso", "Linha editada com sucesso!")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao confirmar edição: {str(e)}")
            
            frame_botoes = ttk.Frame(janela_edicao)
            frame_botoes.pack(fill="x", pady=10)
            
            ttk.Button(frame_botoes, text="Confirmar", 
                      command=confirmar_edicao).pack(side="left", padx=10)
            ttk.Button(frame_botoes, text="Cancelar", 
                      command=janela_edicao.destroy).pack(side="left", padx=10)
            
            # Usar protocolo para evitar problemas com grab_set
            janela_edicao.protocol("WM_DELETE_WINDOW", janela_edicao.destroy)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao criar janela de edição de linha: {str(e)}")
    
    def adicionar_linha(self):
        """Adiciona uma nova linha vazia"""
        try:
            # Adicionar linha vazia ao DataFrame
            nova_linha = {col: "" for col in self.df.columns}
            novo_index = len(self.df)
            self.df.loc[novo_index] = nova_linha
            
            # Atualizar Treeview
            valores = ["" for _ in self.df.columns]
            self.tree.insert("", "end", values=valores, iid=str(novo_index))
            
            messagebox.showinfo("Sucesso", "Nova linha adicionada!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao adicionar linha: {str(e)}")
    
    def remover_linha(self):
        """Remove a linha selecionada"""
        try:
            selecionados = self.tree.selection()
            if not selecionados:
                messagebox.showwarning("Aviso", "Selecione uma linha para remover.")
                return
            
            # Confirmar remoção
            if not messagebox.askyesno("Confirmar", "Tem certeza que deseja remover a(s) linha(s) selecionada(s)?"):
                return
            
            indices_para_remover = []
            for item in selecionados:
                indices_para_remover.append(int(item))
                self.tree.delete(item)
            
            # Remover do DataFrame
            self.df = self.df.drop(indices_para_remover)
            
            # Reindexar DataFrame
            self.df = self.df.reset_index(drop=True)
            
            # Atualizar Treeview com novos índices
            self.atualizar_treeview()
            
            messagebox.showinfo("Sucesso", f"{len(selecionados)} linha(s) removida(s) com sucesso!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao remover linha: {str(e)}")
    
    def salvar_alteracoes(self):
        """Salva as alterações na planilha original"""
        try:
            # Salvar na aba LANÇAMENTOS
            with pd.ExcelWriter(self.arquivo, engine="openpyxl", mode="a", 
                              if_sheet_exists="replace") as writer:
                self.df.to_excel(writer, sheet_name="LANÇAMENTOS", index=False)
            
            # Atualizar também a cópia original
            self.df_original = self.df.copy()
            
            messagebox.showinfo("Sucesso", "Alterações salvas com sucesso!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar alterações: {str(e)}")
    
    def gerar_relatorio(self):
        """Gera relatório a partir dos dados da base de dados"""
        try:
            if self.df.empty:
                messagebox.showwarning("Aviso", "Não há dados para gerar o relatório.")
                return
            
            # Verificar se há dados necessários
            colunas_necessarias = ['Evento', 'Inicio', 'Termino', 'Categoria', 'Total']
            for coluna in colunas_necessarias:
                if coluna not in self.df.columns:
                    messagebox.showerror("Erro", f"Coluna '{coluna}' não encontrada na base de dados.")
                    return
            
            # Obter dados únicos para o cabeçalho
            eventos_unicos = self.df['Evento'].unique()
            if len(eventos_unicos) > 1:
                messagebox.showwarning("Aviso", "Selecione apenas um evento para gerar o relatório.")
                return
            
            evento = eventos_unicos[0]
            primeira_linha = self.df.iloc[0]
            
            # Criar dados do orçamento a partir da base de dados
            dados_orcamento = {
                "CLIENTE:": "",  # Pode ser preenchido manualmente se necessário
                "PROJETO/EVENTO:": str(evento),
                "SOLICITADO POR:": "",
                "TELEFONE:": "",
                "E-MAIL:": "",
                "GERENTE DO PROJETO:": "",
                "RECEBIDO EM:": "",
                "APROVADO EM:": "",
                "DATA DE INICIO": str(primeira_linha['Inicio']),
                "DATA DE TÉRMINO": str(primeira_linha['Termino']),
                "Enviado em ": datetime.now().strftime("%d/%m/%Y")
            }
            
            # Calcular totais por categoria
            try:
                # Converter coluna Total para numérico
                self.df['Total'] = pd.to_numeric(self.df['Total'], errors='coerce').fillna(0)
                totais_categoria = self.df.groupby('Categoria')['Total'].sum()
            except Exception as e:
                print(f"Erro ao calcular totais: {e}")
                totais_categoria = pd.Series()
            
            dados_centro_custo = {
                "ESTRUTURA & CENOGRAFIA": totais_categoria.get("ESTRUTURA & CENOGRAFIA", 0),
                "ATRAÇÕES": totais_categoria.get("ATRAÇÕES", 0),
                "ALIMENTOS E BEBIDAS": totais_categoria.get("ALIMENTOS E BEBIDAS", 0),
                "LOCAÇÃO DE EQUIPAMENTOS": totais_categoria.get("LOCAÇÃO DE EQUIPAMENTOS", 0),
                "SERVIÇOS": totais_categoria.get("SERVIÇOS", 0),
                "EQUIPE/PRODUÇÃO": totais_categoria.get("EQUIPE/PRODUÇÃO", 0),
                "TAXAS/LEGALIZAÇÃO": totais_categoria.get("TAXAS/LEGALIZAÇÃO", 0),
                "DIVULGAÇÃO": totais_categoria.get("DIVULGAÇÃO", 0),
                "OUTROS": totais_categoria.get("OUTROS", 0)
            }
            
            print("Dados para relatório:")
            print("Orçamento:", dados_orcamento)
            print("Centro de custo:", dados_centro_custo)
            
            # Gerar relatório
            self._gerar_relatorio_excel(dados_orcamento, dados_centro_custo, self.df)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório: {str(e)}")
            print(f"Erro detalhado: {e}")
    
    def _gerar_relatorio_excel(self, dados_orcamento, dados_centro_custo, df_itens):
        """Gera o relatório Excel com os dados fornecidos"""
        try:
            arquivo_editavel = xe.cop_sheet(str(self.arquivo))
            wb = load_workbook(arquivo_editavel)
            ws = wb["ENVIO P CLIENTE"]
            df = pd.read_excel(arquivo_editavel, sheet_name="ENVIO P CLIENTE")
            
            # Salvar cabeçalho
            self._salvar_cabecalho_excel(df, ws, dados_orcamento, dados_centro_custo)
            wb.save(arquivo_editavel)
            
            # Processar categorias
            contagem = df_itens.groupby('Categoria').size().to_dict()
            
            categorias_para_processar = [
                "ESTRUTURA & CENOGRAFIA",
                "ATRAÇÕES",
                "ALIMENTOS E BEBIDAS", 
                "LOCAÇÃO DE EQUIPAMENTOS",
                "SERVIÇOS",
                "EQUIPE/PRODUÇÃO",
                "TAXAS/LEGALIZAÇÃO",
                "DIVULGAÇÃO",
                "OUTROS"
            ]
            
            for i, categoria in enumerate(categorias_para_processar):
                if categoria not in contagem or contagem[categoria] == 0:
                    continue
                    
                wb = load_workbook(arquivo_editavel)
                ws = wb["ENVIO P CLIENTE"]
                df = pd.read_excel(arquivo_editavel, sheet_name="ENVIO P CLIENTE")
                
                posicoes = self._encontrar_posicoes_celula(df, categoria)
                if len(posicoes) < 2:
                    print(f"Categoria {categoria} não encontrada na planilha")
                    continue
                    
                linha_categoria, _ = posicoes[1]
                linha_categoria += 2
                
                if i == 0:
                    linha_base = linha_categoria + 1
                    subtotal_line = self._obter_dados_linha(df, ws, 'SUBTOTAL', 0)
                    subtotal_line_ultimo = self._obter_dados_linha(df, ws, 'SUBTOTAL', -1)
                    imposto_line = self._obter_dados_linha(df, ws, 'IMPOSTOS', -1)
                    
                    posicoes_total = self._encontrar_posicoes_celula(df, 'TOTAL GERAL')
                    if not posicoes_total:
                        print("TOTAL GERAL não encontrado")
                        continue
                    linha_total, _ = posicoes_total[-1]
                    linha_total += 2
                    total_line_data = xe.store_line_data(ws, linha_total, 10)
                    ultima_linha = linha_total + 2

                linha_detalhes = linha_categoria + 1
                
                try:
                    ws.merge_cells(start_row=linha_categoria, start_column=1, end_row=linha_categoria, end_column=9)
                except:
                    pass
                    
                xe.copy_line_format(ws, linha_base, linha_detalhes, 10, True)

                linha = linha_categoria + 2
                ws.insert_rows(linha, contagem[categoria])
                wb.save(arquivo_editavel)
                
                wb = load_workbook(arquivo_editavel)
                ws = wb["ENVIO P CLIENTE"]
                
                for linha_idx in range(linha, linha + contagem[categoria]):
                    xe.copy_line_format(ws, linha_base, linha_idx, 10)

                # Preencher dados dos itens
                indice = 0
                itens_categoria = df_itens[df_itens['Categoria'] == categoria]
                
                for _, item in itens_categoria.iterrows():
                    # Encontrar a coluna de descrição (pode ser 'Descrição' ou 'descricao')
                    coluna_descricao = 'Descrição' if 'Descrição' in item else 'descricao'
                    descricao = item[coluna_descricao] if coluna_descricao in item else ''
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 1)
                    celula.value = item['Item']
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 2)
                    celula.value = item['tipo']
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 3)
                    celula.value = descricao
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 5)
                    celula.value = item['Quantidade']
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 6)
                    celula.value = item['Diária']
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 7)
                    celula.value = item['Subtotal']
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 8)
                    celula.value = item['Total']
                    
                    celula = xe.encontrar_celula_editavel(ws, linha + indice, 9)
                    celula.value = item['Incluir']
                    
                    indice += 1

                xe.apply_line_data(ws, linha + contagem[categoria], subtotal_line, True)
                wb.save(arquivo_editavel)

            # Aplicar totais finais
            wb = load_workbook(arquivo_editavel)
            ws = wb["ENVIO P CLIENTE"]
            df = pd.read_excel(arquivo_editavel, sheet_name="ENVIO P CLIENTE")
            
            posicoes_total = self._encontrar_posicoes_celula(df, 'TOTAL GERAL')
            if posicoes_total:
                linha_total, _ = posicoes_total[-1]
                linha_total += 2
                linha_imposto = linha_total - 1
                linha_subtotal = linha_imposto - 1
                ultima_linha = linha_total + 2

                xe.apply_line_data(ws, linha_subtotal, subtotal_line_ultimo, True)
                xe.apply_line_data(ws, linha_imposto, imposto_line, True)
                xe.apply_line_data(ws, linha_total, total_line_data, True)
                
                try:
                    ws.merge_cells(start_row=ultima_linha, start_column=1, end_row=ultima_linha, end_column=9)
                except:
                    pass
                
                wb.save(arquivo_editavel)

            messagebox.showinfo("Sucesso", "Relatório Gerado com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar relatório Excel: {str(e)}")
            print(f"Erro detalhado ao gerar Excel: {e}")
    
    def _salvar_cabecalho_excel(self, df, ws, dados_orcamento, dados_centro_custo):
        """Salva os dados do cabeçalho na planilha Excel"""
        for dado in dados_orcamento:
            posicoes = self._encontrar_posicoes_celula(df, dado)
            if not posicoes:
                continue
                
            linha, coluna = posicoes[0]
            linha += 2
            coluna += 1
            
            celula = ws.cell(row=linha, column=coluna)
            
            if dado == "Enviado em ":
                celula.value = f'Enviado em {dados_orcamento[dado]}'
            elif dado in ["DATA DE INICIO", "DATA DE TÉRMINO", "GERENTE DO PROJETO:", "RECEBIDO EM:", "APROVADO EM:"]:
                celula = ws.cell(row=linha, column=coluna + 2)
                celula.value = dados_orcamento[dado]
            else:
                celula = ws.cell(row=linha, column=coluna + 1)
                celula.value = dados_orcamento[dado]

        for dado in dados_centro_custo:
            posicoes = self._encontrar_posicoes_celula(df, dado)
            if not posicoes:
                continue
                
            linha, coluna = posicoes[0]
            linha += 2
            coluna += 1
            
            celula = ws.cell(row=linha, column=coluna + 6)
            valor = dados_centro_custo[dado]
            if pd.notna(valor) and valor != "":
                try:
                    celula.value = f'R$ {float(valor):.2f}'
                except:
                    celula.value = f'R$ {valor}'
    
    @staticmethod
    def _encontrar_posicoes_celula(df: pd.DataFrame, expressao: str) -> List[Tuple[int, int]]:
        """Encontra posições de células que correspondem à expressão"""
        padrao = f".*{re.escape(expressao)}.*"
        posicoes = []

        for linha in range(len(df)):
            for coluna in range(len(df.columns)):
                valor = str(df.iat[linha, coluna])
                if re.search(padrao, valor):
                    posicoes.append((linha, coluna))

        return posicoes if posicoes else []
    
    def _obter_dados_linha(self, df: pd.DataFrame, ws: Any, regex: str, indice: int) -> Dict[str, List]:
        """Obtém dados de uma linha específica baseada em regex"""
        posicoes = self._encontrar_posicoes_celula(df, regex)
        if not posicoes or len(posicoes) <= indice:
            return {}
        linha, _ = posicoes[indice]
        linha += 2
        return xe.store_line_data(ws, linha, 10)
class BudgetApp:
    """Classe principal da aplicação de orçamento"""
    
    def __init__(self):
        self.dados_orcamento: Dict[str, str] = {}
        self.dados_centro_custo: Dict[str, str] = {}
        self.linhas_categoria: List[Dict] = []
        self.categorias_preenchidas: List[str] = []
        self.arquivo: Optional[Path] = None
        
        self._setup_paths()
        self._setup_ui()
        
    def _setup_paths(self) -> None:
        """Configura os caminhos de arquivo"""
        pasta_atual = Path.cwd()
        self.arquivo = pasta_atual / "ORÇAMENTO DE EVENTOS.xlsx"
        
        if not self.arquivo.exists():
            self.arquivo = Path(self.escolhe_arquivo())

    def _setup_ui(self) -> None:
        """Configura a interface gráfica"""
        self.janela = tk.Tk()
        self.janela.title("Sistema de Orçamento")
        self.janela.geometry("1000x700")

        self._criar_menu_principal()
        self._criar_frames()
        self._criar_widgets_orcamento()
        self._criar_widgets_centro_custo()
        self._criar_widgets_detalhamento()
        self._criar_widgets_base_dados()
        
        # Iniciar com o frame do orçamento visível
        self.frame_centro_custo.pack_forget()
        self.frame_detalhamento.pack_forget()
        self.frame_base_dados.pack_forget()

    def _criar_menu_principal(self):
        """Cria o menu principal da aplicação"""
        menubar = tk.Menu(self.janela)
        self.janela.config(menu=menubar)
        
        menu_arquivo = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Arquivo", menu=menu_arquivo)
        menu_arquivo.add_command(label="Base de Dados", command=self.mostrar_base_dados)
        menu_arquivo.add_separator()
        menu_arquivo.add_command(label="Sair", command=self.janela.quit)
        
        menu_navegacao = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Navegação", menu=menu_navegacao)
        menu_navegacao.add_command(label="Orçamento", command=self.mostrar_orcamento)
        menu_navegacao.add_command(label="Centro de Custo", command=self.mostrar_centro_custo)
        menu_navegacao.add_command(label="Detalhamento", command=self.mostrar_detalhamento_custos)

    def _criar_frames(self) -> None:
        """Cria os frames principais da aplicação"""
        self.frame_orcamento = ttk.Frame(self.janela)
        self.frame_centro_custo = ttk.Frame(self.janela)
        self.frame_detalhamento = ttk.Frame(self.janela)
        self.frame_base_dados = ttk.Frame(self.janela)
        
        self.frame_orcamento.pack(fill="both", expand=True)

    def _criar_widgets_base_dados(self):
        """Cria a interface da base de dados"""
        self.viewer_base_dados = DatabaseViewer(self.frame_base_dados, self.arquivo)
        self.viewer_base_dados.frame.pack(fill="both", expand=True)
        
        # Botão para voltar
        frame_controles = ttk.Frame(self.frame_base_dados)
        frame_controles.pack(fill="x", pady=10)
        
        ttk.Button(frame_controles, text="Voltar para Orçamento", 
                  command=self.mostrar_orcamento).pack(side="left", padx=10)

    def mostrar_base_dados(self):
        """Mostra a tela da base de dados"""
        self.frame_orcamento.pack_forget()
        self.frame_centro_custo.pack_forget()
        self.frame_detalhamento.pack_forget()
        self.frame_base_dados.pack(fill="both", expand=True)
        
        # Recarregar dados ao mostrar
        self.viewer_base_dados.carregar_dados()

    def _criar_scrollable_frame(self, parent: tk.Widget) -> Tuple[tk.Canvas, ttk.Frame]:
        """Cria um frame com scrollbar"""
        container = ttk.Frame(parent)
        container.pack(fill="both", expand=True)

        canvas = tk.Canvas(container)
        v_scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        h_scrollbar = ttk.Scrollbar(container, orient="horizontal", command=canvas.xview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")

        return canvas, scrollable_frame

    def _criar_widgets_orcamento(self) -> None:
        """Cria os widgets do frame de orçamento"""
        canvas, scrollable_frame = self._criar_scrollable_frame(self.frame_orcamento)
        canvas.bind("<MouseWheel>", self._on_mousewheel)

        scrollable_frame.columnconfigure(1, weight=1)

        # Campos do orçamento
        campos_orcamento = [
            ("CLIENTE:", "entry_cliente"),
            ("PROJETO/EVENTO:", "entry_evento"),
            ("SOLICITADO POR:", "entry_solicitado"),
            ("TELEFONE:", "entry_telefone"),
            ("E-MAIL:", "entry_email"),
            ("GERENTE DO PROJETO:", "entry_gerente"),
            ("RECEBIDO EM:", "entry_recebido"),
            ("APROVADO EM:", "entry_aprovado"),
            ("DATA DE INICIO:", "entry_inicio"),
            ("DATA DE TÉRMINO:", "entry_termino"),
            ("Enviado em ", "entry_enviado")
        ]

        self.entries_orcamento = {}
        for i, (label_text, entry_name) in enumerate(campos_orcamento):
            ttk.Label(scrollable_frame, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky="e")
            entry = ttk.Entry(scrollable_frame, width=40)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
            self.entries_orcamento[entry_name] = entry

        # Botões do orçamento
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.grid(row=len(campos_orcamento), column=0, columnspan=2, pady=20)

        ttk.Button(button_frame, text="Salvar e Continuar", command=self.salvar_orcamento).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Limpar", command=self.limpar_campos).pack(side="left", padx=10)

    def _criar_widgets_centro_custo(self) -> None:
        """Cria os widgets do frame de centro de custo"""
        canvas, scrollable_frame = self._criar_scrollable_frame(self.frame_centro_custo)
        canvas.bind("<MouseWheel>", self._on_mousewheel_centro_custo)

        scrollable_frame.columnconfigure(1, weight=1)

        # Categorias do centro de custo
        categorias = [
            "ESTRUTURA & CENOGRAFIA R$",
            "ATRAÇÕES R$",
            "ALIMENTOS E BEBIDAS R$",
            "LOCAÇÃO DE EQUIPAMENTOS R$",
            "SERVIÇOS R$",
            "EQUIPE/PRODUÇÃO R$",
            "TAXAS/LEGALIZAÇÃO R$",
            "DIVULGAÇÃO R$",
            "OUTROS R$"
        ]

        self.entries_centro_custo = {}
        for i, categoria in enumerate(categorias):
            ttk.Label(scrollable_frame, text=categoria, font=('Arial', 10, 'bold')).grid(
                row=i, column=0, padx=5, pady=8, sticky="e"
            )
            entry = ttk.Entry(scrollable_frame, width=20)
            entry.grid(row=i, column=1, padx=5, pady=8, sticky="ew")
            self.entries_centro_custo[categoria] = entry

        # Botões do centro de custo
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.grid(row=len(categorias), column=0, columnspan=2, pady=20)

        ttk.Button(button_frame, text="Salvar e Continuar", command=self.salvar_centro_custo).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Voltar", command=self.voltar_orcamento).pack(side="left", padx=10)

    def _criar_widgets_detalhamento(self) -> None:
        """Cria os widgets do frame de detalhamento"""
        # Notebook para as abas de categorias
        self.notebook = ttk.Notebook(self.frame_detalhamento)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        # Botões do frame de detalhamento
        button_frame = ttk.Frame(self.frame_detalhamento)
        button_frame.pack(fill="x", pady=10)

        ttk.Button(button_frame, text="Novo Orçamento", command=self.finalizar_orcamento).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Gerar Relatório", command=self.salvar_relatorio).pack(side="left", padx=10)
        ttk.Button(button_frame, text="Voltar", command=self.voltar_centro_custo).pack(side="left", padx=10)

    # Métodos de negócio
    def salvar_orcamento(self) -> None:
        """Salva os dados do orçamento"""
        self.dados_orcamento = {
            "CLIENTE:": self.entries_orcamento["entry_cliente"].get(),
            "PROJETO/EVENTO:": self.entries_orcamento["entry_evento"].get(),
            "SOLICITADO POR:": self.entries_orcamento["entry_solicitado"].get(),
            "TELEFONE:": self.entries_orcamento["entry_telefone"].get(),
            "E-MAIL:": self.entries_orcamento["entry_email"].get(),
            "GERENTE DO PROJETO:": self.entries_orcamento["entry_gerente"].get(),
            "RECEBIDO EM:": self.entries_orcamento["entry_recebido"].get(),
            "APROVADO EM:": self.entries_orcamento["entry_aprovado"].get(),
            "DATA DE INICIO": self.entries_orcamento["entry_inicio"].get(),
            "DATA DE TÉRMINO": self.entries_orcamento["entry_termino"].get(),
            "Enviado em ": self.entries_orcamento["entry_enviado"].get()
        }

        if not self.dados_orcamento["CLIENTE:"] or not self.dados_orcamento["PROJETO/EVENTO:"]:
            messagebox.showerror("Erro", "Preencha pelo menos o nome do Cliente e o Projeto/Evento.")
            return

        messagebox.showinfo("Sucesso", "Orçamento salvo com sucesso! Agora preencha o Orçamento Resumido.")
        self.mostrar_centro_custo()

    def salvar_centro_custo(self) -> None:
        """Salva os dados do centro de custo"""
        self.dados_centro_custo = {
            "ESTRUTURA & CENOGRAFIA": self.entries_centro_custo["ESTRUTURA & CENOGRAFIA R$"].get(),
            "ATRAÇÕES": self.entries_centro_custo["ATRAÇÕES R$"].get(),
            "ALIMENTOS E BEBIDAS": self.entries_centro_custo["ALIMENTOS E BEBIDAS R$"].get(),
            "LOCAÇÃO DE EQUIPAMENTOS": self.entries_centro_custo["LOCAÇÃO DE EQUIPAMENTOS R$"].get(),
            "SERVIÇOS": self.entries_centro_custo["SERVIÇOS R$"].get(),
            "EQUIPE/PRODUÇÃO": self.entries_centro_custo["EQUIPE/PRODUÇÃO R$"].get(),
            "TAXAS/LEGALIZAÇÃO": self.entries_centro_custo["TAXAS/LEGALIZAÇÃO R$"].get(),
            "DIVULGAÇÃO": self.entries_centro_custo["DIVULGAÇÃO R$"].get(),
            "OUTROS": self.entries_centro_custo["OUTROS R$"].get()
        }

        if not any(self.dados_centro_custo.values()):
            messagebox.showwarning("Aviso", "Nenhum valor foi preenchido no Orçamento Resumido.")
            return

        messagebox.showinfo("Sucesso", "Orçamento Resumido salvo! Agora detalhe os itens por categoria.")
        self.mostrar_detalhamento_custos()

    def mostrar_centro_custo(self) -> None:
        """Mostra o frame de centro de custo"""
        self.frame_orcamento.pack_forget()
        self.frame_centro_custo.pack(fill="both", expand=True)

    def voltar_orcamento(self) -> None:
        """Volta para o frame de orçamento"""
        self.frame_centro_custo.pack_forget()
        self.frame_orcamento.pack(fill="both", expand=True)

    def mostrar_detalhamento_custos(self) -> None:
        """Mostra o frame de detalhamento de custos"""
        self.frame_centro_custo.pack_forget()
        self.frame_detalhamento.pack(fill="both", expand=True)

        # Limpar abas anteriores
        for aba in self.notebook.tabs():
            self.notebook.forget(aba)

        # Criar abas apenas para categorias preenchidas
        self.categorias_preenchidas = [
            categoria for categoria, valor in self.dados_centro_custo.items() if valor
        ]

        # Mapear nomes das categorias para exibição
        nomes_categorias = {
            "ESTRUTURA & CENOGRAFIA": "ESTRUTURA & CENOGRAFIA",
            "ATRAÇÕES": "ATRAÇÕES",
            "ALIMENTOS E BEBIDAS": "ALIMENTOS E BEBIDAS",
            "LOCAÇÃO DE EQUIPAMENTOS": "LOCAÇÃO DE EQUIPAMENTOS",
            "SERVIÇOS": "SERVIÇOS",
            "EQUIPE/PRODUÇÃO": "EQUIPE/PRODUÇÃO",
            "TAXAS/LEGALIZAÇÃO": "TAXAS/LEGALIZAÇÃO",
            "DIVULGAÇÃO": "DIVULGAÇÃO",
            "OUTROS": "OUTROS"
        }

        # Criar uma aba para cada categoria preenchida
        for categoria in self.categorias_preenchidas:
            self._criar_aba_categoria(nomes_categorias[categoria])

    def _criar_aba_categoria(self, nome_categoria: str) -> None:
        """Cria uma aba para detalhamento de categoria"""
        frame_aba = ttk.Frame(self.notebook)
        self.notebook.add(frame_aba, text=nome_categoria)

        canvas, scrollable_frame = self._criar_scrollable_frame(frame_aba)

        # Cabeçalho da tabela
        colunas = ["Item", "Tipo", "Descrição", "Qtde", "Diária", "Subtotal", "Total", "Incluir no Orçamento?"]
        pesos_colunas = [2, 2, 3, 2, 2, 2, 2, 2]
        
        for i, peso in enumerate(pesos_colunas):
            scrollable_frame.columnconfigure(i, weight=peso)

        for i, coluna in enumerate(colunas):
            ttk.Label(scrollable_frame, text=coluna, font=('Arial', 9, 'bold')).grid(
                row=0, column=i, padx=2, pady=5, sticky="ew"
            )

        # Frame para os campos de entrada
        frame_campos = ttk.Frame(scrollable_frame)
        frame_campos.grid(row=1, column=0, columnspan=len(colunas), sticky="ew")

        for i, peso in enumerate(pesos_colunas):
            frame_campos.columnconfigure(i, weight=peso)

        # Botões da aba
        btn_frame = ttk.Frame(scrollable_frame)
        btn_frame.grid(row=2, column=0, columnspan=4, pady=10, sticky="w")

        ttk.Button(btn_frame, text="+ Adicionar Linha", 
                  command=lambda: self._adicionar_linha(frame_campos, nome_categoria, pesos_colunas, colunas)).pack(side="left")
        ttk.Button(btn_frame, text=f"Salvar {nome_categoria}", 
                  command=lambda: self._salvar_categoria(nome_categoria)).pack(side="left", padx=10)

        # Adicionar primeira linha automaticamente
        self._adicionar_linha(frame_campos, nome_categoria, pesos_colunas, colunas)

    def _adicionar_linha(self, parent: tk.Widget, categoria: str, pesos_colunas: List[int], colunas: List[str]) -> None:
        """Adiciona uma nova linha de entrada na categoria"""
        row_index = len(self.linhas_categoria) + 1

        frame_linha = ttk.Frame(parent)
        frame_linha.grid(row=row_index, column=0, columnspan=len(colunas), sticky="ew", pady=1)

        for i, peso in enumerate(pesos_colunas):
            frame_linha.columnconfigure(i, weight=peso)

        # Campos da linha
        entry_item = ttk.Entry(frame_linha)
        entry_item.grid(row=0, column=0, padx=2, sticky="ew")

        entry_tipo = ttk.Entry(frame_linha)
        entry_tipo.grid(row=0, column=1, padx=2, sticky="ew")

        entry_descricao = ttk.Entry(frame_linha)
        entry_descricao.grid(row=0, column=2, padx=2, sticky="ew")

        entry_qtde = ttk.Entry(frame_linha)
        entry_qtde.grid(row=0, column=3, padx=2, sticky="ew")
        entry_qtde.insert(0, "1")

        entry_diaria = ttk.Entry(frame_linha)
        entry_diaria.grid(row=0, column=4, padx=2, sticky="ew")

        entry_subtotal = ttk.Entry(frame_linha)
        entry_subtotal.grid(row=0, column=5, padx=2, sticky="ew")

        entry_total = ttk.Entry(frame_linha)
        entry_total.grid(row=0, column=6, padx=2, sticky="ew")

        var_incluir = tk.BooleanVar(value=True)
        check_incluir = ttk.Checkbutton(frame_linha, variable=var_incluir)
        check_incluir.grid(row=0, column=7, padx=2)

        for i in range(len(colunas)):
            frame_linha.columnconfigure(i, weight=1)

        linha_data = {
            "frame": frame_linha,
            "entries": {
                "categoria": categoria,
                "item": entry_item,
                "tipo": entry_tipo,
                "descricao": entry_descricao,
                "qtde": entry_qtde,
                "diaria": entry_diaria,
                "subtotal": entry_subtotal,
                "total": entry_total,
                "incluir": var_incluir
            }
        }

        self.linhas_categoria.append(linha_data)

    def _salvar_categoria(self, nome_categoria: str) -> None:
        """Salva os dados da categoria específica"""
        itens_categoria = []
        
        for linha in self.linhas_categoria:
            if linha["entries"]["categoria"] == nome_categoria:
                incluir = "Sim" if linha["entries"]["incluir"].get() else "Não"
                item_data = {
                    "categoria": nome_categoria,
                    "item": linha["entries"]["item"].get(),
                    "tipo": linha["entries"]["tipo"].get(),
                    "descricao": linha["entries"]["descricao"].get(),
                    "qtde": linha["entries"]["qtde"].get(),
                    "diaria": linha["entries"]["diaria"].get(),
                    "subtotal": linha["entries"]["subtotal"].get(),
                    "total": linha["entries"]["total"].get(),
                    "incluir": incluir
                }
                itens_categoria.append(item_data)

        dados = []
        for item in itens_categoria:
            dados.append({
                "Inicio": self.dados_orcamento['DATA DE INICIO'],
                "Termino": self.dados_orcamento['DATA DE TÉRMINO'],
                "Evento": self.dados_orcamento['PROJETO/EVENTO:'],
                "Categoria": item['categoria'],
                "Item": item['item'],
                "tipo": item['tipo'],
                "Descrição": item['descricao'],
                "Quantidade": item['qtde'],
                "Diária": item['diaria'],
                "Subtotal": item['subtotal'],
                "Total": item['total'],
                "Incluir": item['incluir']
            })

        lancamento = "LANÇAMENTOS"
        
        try:
            # Carregar dados existentes
            df_existente = pd.read_excel(self.arquivo, sheet_name=lancamento)
            
            # Remover dados do evento atual antes de adicionar os novos
            evento_atual = self.dados_orcamento['PROJETO/EVENTO:']
            df_existente = df_existente[df_existente['Evento'] != evento_atual]
            
            # Combinar com novos dados
            df_final = pd.concat([df_existente, pd.DataFrame(dados)], ignore_index=True)
        except (FileNotFoundError, ValueError):
            # Se a aba não existir, usar apenas os novos dados
            df_final = pd.DataFrame(dados)

        with pd.ExcelWriter(self.arquivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_final.to_excel(writer, sheet_name=lancamento, index=False)

        messagebox.showinfo("Sucesso", f"Itens da categoria {nome_categoria} salvos com sucesso!")
    # Métodos utilitários
    @staticmethod
    def _encontrar_posicoes_celula(df: pd.DataFrame, expressao: str) -> List[Tuple[int, int]]:
        """Encontra posições de células que correspondem à expressão"""
        padrao = f".*{re.escape(expressao)}.*"
        posicoes = []

        for linha in range(len(df)):
            for coluna in range(len(df.columns)):
                valor = str(df.iat[linha, coluna])
                if re.search(padrao, valor):
                    posicoes.append((linha, coluna))

        return posicoes if posicoes else []

    def _salvar_cabecalho(self, df: pd.DataFrame, ws: Any) -> None:
        """Salva os dados do cabeçalho na planilha"""
        for dado in self.dados_orcamento:
            posicoes = self._encontrar_posicoes_celula(df, dado)
            if not posicoes:
                continue
                
            linha, coluna = posicoes[0]
            linha += 2
            coluna += 1
            
            celula = ws.cell(row=linha, column=coluna)
            
            if dado == "Enviado em ":
                celula.value = f'Enviado em {self.dados_orcamento[dado]}'
            elif dado in ["DATA DE INICIO", "DATA DE TÉRMINO", "GERENTE DO PROJETO:", "RECEBIDO EM:", "APROVADO EM:"]:
                celula = ws.cell(row=linha, column=coluna + 2)
                celula.value = self.dados_orcamento[dado]
            else:
                celula = ws.cell(row=linha, column=coluna + 1)
                celula.value = self.dados_orcamento[dado]

        for dado in self.dados_centro_custo:
            posicoes = self._encontrar_posicoes_celula(df, dado)
            if not posicoes:
                continue
                
            linha, coluna = posicoes[0]
            linha += 2
            coluna += 1
            
            celula = ws.cell(row=linha, column=coluna + 6)
            celula.value = f'R$ {self.dados_centro_custo[dado]}'

    def _obter_dados_linha(self, df: pd.DataFrame, ws: Any, regex: str, indice: int) -> Dict[str, List]:
        """Obtém dados de uma linha específica baseada em regex"""
        posicoes = self._encontrar_posicoes_celula(df, regex)
        linha, _ = posicoes[indice]
        linha += 2
        return xe.store_line_data(ws, linha, 10)

    def salvar_relatorio(self) -> None:
        """Salva o relatório completo"""
        try:
            arquivo_editavel = xe.cop_sheet(str(self.arquivo))
            wb = load_workbook(arquivo_editavel)
            ws = wb["ENVIO P CLIENTE"]
            df = pd.read_excel(arquivo_editavel, sheet_name="ENVIO P CLIENTE")
            
            self._salvar_cabecalho(df, ws)
            wb.save(arquivo_editavel)

            # Processar categorias
            contagem = {}
            for linha in self.linhas_categoria:
                categoria = linha['entries']['categoria']
                contagem[categoria] = contagem.get(categoria, 0) + 1

            nomes_categorias = [
                "ESTRUTURA & CENOGRAFIA",
                "ATRAÇÕES",
                "ALIMENTOS E BEBIDAS",
                "LOCAÇÃO DE EQUIPAMENTOS",
                "SERVIÇOS",
                "EQUIPE/PRODUÇÃO",
                "TAXAS/LEGALIZAÇÃO",
                "DIVULGAÇÃO",
                "OUTROS"
                ]
            for i, cat in enumerate(nomes_categorias):
                try:
                    if nomes_categorias[i] == contagem[cat]:
                        continue
                except:
                    contagem[cat] = 1
            print(contagem)
            for i, categoria in enumerate(nomes_categorias):
                wb = load_workbook(arquivo_editavel)
                ws = wb["ENVIO P CLIENTE"]
                df = pd.read_excel(arquivo_editavel, sheet_name="ENVIO P CLIENTE")
                
                posicoes = self._encontrar_posicoes_celula(df, categoria)
                if not posicoes:
                    continue
                linha_categoria, _ = posicoes[1]
                linha_categoria += 2
                
                if i == 0:
                    linha_base = linha_categoria + 1
                    subtotal_line = self._obter_dados_linha(df, ws, 'SUBTOTAL', 0)
                    subtotal_line_ultimo = self._obter_dados_linha(df, ws, 'SUBTOTAL', -1)
                    imposto_line = self._obter_dados_linha(df, ws, 'IMPOSTOS', -1)
                    
                    posicoes_total = self._encontrar_posicoes_celula(df, 'TOTAL GERAL')
                    linha_total, _ = posicoes_total[-1]
                    linha_total += 2
                    total_line_data = xe.store_line_data(ws, linha_total, 10)
                    ultima_linha = linha_total + 2

                linha_detalhes = linha_categoria + 1
                ws.merge_cells(start_row=linha_categoria, start_column=1, end_row=linha_categoria, end_column=9)
                xe.copy_line_format(ws, linha_base, linha_detalhes, 10, True)

                linha = linha_categoria + 2
                ws.insert_rows(linha, contagem[categoria])
                wb.save(arquivo_editavel)
                
                wb = load_workbook(arquivo_editavel)
                ws = wb["ENVIO P CLIENTE"]
                
                for linha_idx in range(linha, linha + contagem[categoria]):
                    xe.copy_line_format(ws, linha_base, linha_idx, 10)

                indice = 0
                for linha_data in self.linhas_categoria:
                    if linha_data["entries"]["categoria"] == categoria:
                        incluir = 'Sim' if linha_data["entries"]["incluir"].get() else 'Não'
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 1)
                        celula.value = linha_data["entries"]["item"].get()
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 2)
                        celula.value = linha_data["entries"]["tipo"].get()
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 3)
                        celula.value = linha_data["entries"]["descricao"].get()
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 5)
                        celula.value = linha_data["entries"]["qtde"].get()
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 6)
                        celula.value = linha_data["entries"]["diaria"].get()
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 7)
                        celula.value = linha_data["entries"]["subtotal"].get()
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 8)
                        celula.value = linha_data["entries"]["total"].get()
                        
                        celula = xe.encontrar_celula_editavel(ws, linha + indice, 9)
                        celula.value = incluir
                        
                        indice += 1
                    else:
                        continue;

                xe.apply_line_data(ws, linha + contagem[categoria], subtotal_line, True)
                wb.save(arquivo_editavel)

            # Aplicar totais finais
            df = pd.read_excel(arquivo_editavel, sheet_name="ENVIO P CLIENTE")
            posicoes_total = self._encontrar_posicoes_celula(df, 'TOTAL GERAL')
            linha_total, _ = posicoes_total[-1]
            linha_total += 2
            linha_imposto = linha_total - 1
            linha_subtotal = linha_imposto - 1
            ultima_linha = linha_total + 2

            xe.apply_line_data(ws, linha_subtotal, subtotal_line_ultimo, True)
            xe.apply_line_data(ws, linha_imposto, imposto_line, True)
            xe.apply_line_data(ws, linha_total, total_line_data, True)
            ws.merge_cells(start_row=ultima_linha, start_column=1, end_row=ultima_linha, end_column=9)
            wb.save(arquivo_editavel)

            messagebox.showinfo("Sucesso", "Relatório Gerado com sucesso!")

        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o relatório: {str(e)}")
            print(e)

    @staticmethod
    def escolhe_arquivo() -> str:
        """Abre diálogo para escolher arquivo"""
        root = tk.Tk()
        root.withdraw()
        caminho_arquivo = filedialog.askopenfilename(
            title="Selecione a planilha de orçamentos",
            filetypes=(("Arquivos Excel", "*.xlsx"), ("Todos os arquivos", "*.*"))
        )
        root.destroy()
        return caminho_arquivo

    def voltar_centro_custo(self) -> None:
        """Volta para o centro de custo"""
        self.frame_detalhamento.pack_forget()
        self.frame_centro_custo.pack(fill="both", expand=True)

    def finalizar_orcamento(self) -> None:
        """Finaliza o orçamento"""
        try:
            # Limpar dados do evento atual da planilha LANÇAMENTOS
            evento_atual = self.dados_orcamento['PROJETO/EVENTO:']
            
            try:
                df_existente = pd.read_excel(self.arquivo, sheet_name="LANÇAMENTOS")
                # Manter apenas dados de outros eventos
                df_final = df_existente[df_existente['Evento'] != evento_atual]
                
                with pd.ExcelWriter(self.arquivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    df_final.to_excel(writer, sheet_name="LANÇAMENTOS", index=False)
            except (FileNotFoundError, ValueError):
                pass  # Se não existir a planilha, não faz nada
                
        except Exception as e:
            print(f"Erro ao limpar dados do evento: {e}")
        
        messagebox.showinfo("Sucesso", "Orçamento completo salvo com sucesso!")
        self.limpar_campos()
        self.resetar_orcamento()

    def resetar_orcamento(self) -> None:
        """Reseta para o início do orçamento"""
        self.frame_detalhamento.pack_forget()
        self.frame_orcamento.pack(fill="both", expand=True)

    def limpar_campos(self) -> None:
        """Limpa todos os campos do formulário"""
        # Limpar campos do orçamento
        for entry in self.entries_orcamento.values():
            entry.delete(0, tk.END)

        # Limpar campos do centro de custo
        for entry in self.entries_centro_custo.values():
            entry.delete(0, tk.END)

        # Limpar dados em memória
        self.dados_orcamento.clear()
        self.dados_centro_custo.clear()
        self.linhas_categoria.clear()
        self.categorias_preenchidas.clear()
        
        # Limpar abas do notebook se existirem
        try:
            for aba in self.notebook.tabs():
                self.notebook.forget(aba)
        except:
            pass
    # Handlers de eventos
    def _on_mousewheel(self, event: tk.Event) -> None:
        """Handler para scroll do mouse"""
        self._scroll_canvas(event, "canvas_orcamento")

    def _on_mousewheel_centro_custo(self, event: tk.Event) -> None:
        """Handler para scroll do mouse no centro de custo"""
        self._scroll_canvas(event, "canvas_centro_custo")

    def _scroll_canvas(self, event: tk.Event, canvas_name: str) -> None:
        """Executa scroll no canvas"""
        canvas = getattr(self, canvas_name, None)
        if canvas:
            canvas.yview_scroll(-1 * int(event.delta / 120), "units")


    def mostrar_orcamento(self):
        """Mostra o frame de orçamento"""
        self.frame_centro_custo.pack_forget()
        self.frame_detalhamento.pack_forget()
        self.frame_base_dados.pack_forget()
        self.frame_orcamento.pack(fill="both", expand=True)

    def mostrar_centro_custo(self):
        """Mostra o frame de centro de custo"""
        self.frame_orcamento.pack_forget()
        self.frame_detalhamento.pack_forget()
        self.frame_base_dados.pack_forget()
        self.frame_centro_custo.pack(fill="both", expand=True)

    def mostrar_detalhamento_custos(self) -> None:
        """Mostra o frame de detalhamento de custos"""
        self.frame_centro_custo.pack_forget()
        self.frame_detalhamento.pack(fill="both", expand=True)

        # Limpar dados anteriores das linhas de categoria
        self.linhas_categoria.clear()

        # Limpar abas anteriores
        for aba in self.notebook.tabs():
            self.notebook.forget(aba)

        # Criar abas apenas para categorias preenchidas
        self.categorias_preenchidas = [
            categoria for categoria, valor in self.dados_centro_custo.items() 
            if valor and str(valor).strip()
        ]

        print(f"Categorias preenchidas: {self.categorias_preenchidas}")

        # Mapear nomes das categorias para exibição
        nomes_categorias = {
            "ESTRUTURA & CENOGRAFIA": "ESTRUTURA & CENOGRAFIA",
            "ATRAÇÕES": "ATRAÇÕES",
            "ALIMENTOS E BEBIDAS": "ALIMENTOS E BEBIDAS",
            "LOCAÇÃO DE EQUIPAMENTOS": "LOCAÇÃO DE EQUIPAMENTOS",
            "SERVIÇOS": "SERVIÇOS",
            "EQUIPE/PRODUÇÃO": "EQUIPE/PRODUÇÃO",
            "TAXAS/LEGALIZAÇÃO": "TAXAS/LEGALIZAÇÃO",
            "DIVULGAÇÃO": "DIVULGAÇÃO",
            "OUTROS": "OUTROS"
        }

        # Criar uma aba para cada categoria preenchida
        for categoria in self.categorias_preenchidas:
            if categoria in nomes_categorias:
                self._criar_aba_categoria(nomes_categorias[categoria])
    def executar(self) -> None:
        """Executa a aplicação"""
        self.janela.mainloop()


if __name__ == "__main__":
    app = BudgetApp()
    app.executar()
