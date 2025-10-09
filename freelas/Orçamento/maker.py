import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import re
from pathlib import Path
import xlsxEditor as xe
from openpyxl import load_workbook

# Variáveis globais para armazenar os dados
dados_orcamento = {}
dados_centro_custo = {}
linhas_categoria = []
categorias_preenchidas = []

def salvar_orcamento():
    global dados_orcamento
    dados_orcamento = {
        "CLIENTE:": entry_cliente.get(),
        "PROJETO/EVENTO:": entry_evento.get(),
        "SOLICITADO POR:": entry_solicitado.get(),
        "TELEFONE:": entry_telefone.get(),
        "E-MAIL:": entry_email.get(),
        "GERENTE DO PROJETO:": entry_gerente.get(),
        "RECEBIDO EM:": entry_recebido.get(),
        "APROVADO EM:": entry_aprovado.get(),
        "DATA DE INICIO": entry_inicio.get(),
        "DATA DE TÉRMINO": entry_termino.get(),
        "Enviado em ": entry_enviado.get()
    }

    if not dados_orcamento["CLIENTE:"] or not dados_orcamento["PROJETO/EVENTO:"]:
        messagebox.showerror("Erro", "Preencha pelo menos o nome do Cliente e o Projeto/Evento.")
        return

    messagebox.showinfo("Sucesso", "Orçamento salvo com sucesso! Agora preencha o Orçamento Resumido.")
    mostrar_centro_custo()

def mostrar_centro_custo():
    frame_orcamento.pack_forget()
    frame_centro_custo.pack(fill="both", expand=True)

def voltar_orcamento():
    frame_centro_custo.pack_forget()
    frame_orcamento.pack(fill="both", expand=True)

def resetar_orcamento():
    frame_detalhamento.pack_forget()
    frame_orcamento.pack(fill="both", expand=True)

def salvar_centro_custo():
    global dados_centro_custo
    dados_centro_custo = {
        "ESTRUTURA & CENOGRAFIA": entry_estrutura.get(),
        "ATRAÇÕES": entry_atracoes.get(),
        "ALIMENTOS E BEBIDAS": entry_alimentos.get(),
        "LOCAÇÃO DE EQUIPAMENTOS": entry_equipamentos.get(),
        "SERVIÇOS": entry_servicos.get(),
        "EQUIPE/PRODUÇÃO": entry_equipe.get(),
        "TAXAS/LEGALIZAÇÃO": entry_taxas.get(),
        "DIVULGAÇÃO": entry_divulgacao.get(),
        "OUTROS": entry_outros.get()
    }

    valores_preenchidos = any(dados_centro_custo.values())
    if not valores_preenchidos:
        messagebox.showwarning("Aviso", "Nenhum valor foi preenchido no Orçamento Resumido.")
        return

    messagebox.showinfo("Sucesso", "Orçamento Resumido salvo! Agora detalhe os itens por categoria.")
    mostrar_detalhamento_custos()

def mostrar_detalhamento_custos():
    frame_centro_custo.pack_forget()
    frame_detalhamento.pack(fill="both", expand=True)
    
    # Limpar abas anteriores
    for aba in notebook.tabs():
        notebook.forget(aba)
    
    # Criar abas apenas para categorias preenchidas
    for categoria, valor in dados_centro_custo.items():
        categorias_preenchidas.append(categoria)
    
    # Mapear nomes das categorias para exibição
    nomes_categorias = {
        "ESTRUTURA & CENOGRAFIA": "ESTRUTURA & CENOGRAFIA",
        "ATRAÇÕES": "ATRAÇÕES",
        "ALIMENTOS E BEBIDAS": "ALIMENTOS E BEBIDAS",
        "LOCAÇÃO DE EQUIPAMENTOS": "LOCAÇÃO DE EQUIPAMENTOS",
        "SERVIÇOS": "SERVIÇOS",
        "EQUIPE/PRODUÇÃO": "EQUIPE/PRODUÇÃO",
        "TAXAS/LEGALIZAÇÃO": "TAXAS/LEGALIZAÇÃO",
        "DIVULGAÇÃO": "DIVULGAÇÃO",
        "OUTROS": "OUTROS"
    }
    
    # Criar uma aba para cada categoria preenchida
    for categoria in categorias_preenchidas:
        criar_aba_categoria(nomes_categorias[categoria])

def acha_lc(df, expressao):
    padrao = f".*{re.escape(expressao)}.*"  # evita erro com caracteres especiais

    posicoes = []

    # percorre todas as células do DataFrame
    for linha in range(len(df)):
        for coluna in range(len(df.columns)):
            valor = str(df.iat[linha, coluna])  # pega valor da célula como string
            if re.search(padrao, valor):
                posicoes.append((linha, coluna))

    if posicoes:
        return posicoes
    else:
        print("Expressão não encontrada")
        return []
def salvar_cabecalho(df,ws):
    for dados in dados_orcamento:
        posicoes = acha_lc(df,dados)
        if not posicoes:
            pass
        linha, coluna = posicoes[0]
        linha+=2
        coluna+=1
        if dados == "Enviado em ":
            celula = ws.cell(row=linha,column=coluna)
            celula.value = f'Enviado em {dados_orcamento[dados]}'
        elif dados == "DATA DE INICIO" or dados == "DATA DE TÉRMINO" or dados == "GERENTE DO PROJETO:" or dados == "RECEBIDO EM:" or dados == "APROVADO EM:":
            celula = ws.cell(row=linha,column=coluna+2)
            celula.value = dados_orcamento[dados]
        else:
            celula = ws.cell(row=linha,column=coluna+1)
            celula.value = dados_orcamento[dados]
       
    for dados in dados_centro_custo:
        posicoes = acha_lc(df,dados)
        if not posicoes:
            pass
        linha, coluna = posicoes[0]
        linha+=2
        coluna+=1
        celula = ws.cell(row=linha,column=coluna+6)
        celula.value = f'R$ {dados_centro_custo[dados]}'

def store_by_regex(df,ws,regex,indice):
            posicoes = acha_lc(df,regex)
            line,_ = posicoes[indice]
            line+=2
            data_line = xe.store_line_data(ws,line,10)
            return data_line

def salvar_relatorio():
    arquivo_editavel = xe.cop_sheet(str(arquivo))
    wb = load_workbook(arquivo_editavel)
    ws = wb["ENVIO P CLIENTE"]
    df = pd.read_excel(arquivo_editavel,"ENVIO P CLIENTE")
    salvar_cabecalho(df,ws)
    global linhas_categoria
    contagem = {}
    for linha in linhas_categoria:
        categoria = linha['entries']['categoria']
        if categoria in contagem:
            contagem[categoria] += 1
        else:
            contagem[categoria] = 1
    for i, categoria in enumerate(categorias_preenchidas):
        wb = load_workbook(arquivo_editavel)
        ws = wb["ENVIO P CLIENTE"]
        df = pd.read_excel(arquivo_editavel,"ENVIO P CLIENTE")
        posicoes = acha_lc(df,categoria)
        line_category,_ = posicoes[1]
        line_category +=2
        if i == 0:
            base_line = line_category + 1
            subtotal_line = store_by_regex(df,ws,'SUBTOTAL',0)
            subtotal_line_ultimo = store_by_regex(df,ws,'SUBTOTAL',-1)
            imposto_line = store_by_regex(df,ws,'IMPOSTOS',-1)
            posicoes = acha_lc(df,'TOTAL GERAL')
            total_line,_ = posicoes[-1]
            total_line+=2
            total_line_data = xe.store_line_data(ws,total_line,10)
            lastline = total_line+2
        details_line = line_category + 1

        ws.merge_cells(start_row=line_category, start_column=1, end_row=line_category, end_column=9)
        xe.copy_line_format(ws,base_line,details_line,10,True)
        print(categoria)
        if not posicoes:
            pass
        line = line_category+2
        ws.insert_rows(idx=line, amount=contagem[categoria])
        wb.save(arquivo_editavel)
        wb = load_workbook(arquivo_editavel)
        ws = wb["ENVIO P CLIENTE"]
        for linha in range(line, line+contagem[categoria]):
            xe.copy_line_format(ws,base_line,linha,10)
        xe.apply_line_data(ws,line+contagem[categoria],subtotal_line,True)
    df = pd.read_excel(arquivo_editavel,"ENVIO P CLIENTE")
    posicoes = acha_lc(df,'TOTAL GERAL')
    tot_line,_ = posicoes[-1]
    tot_line+=2
    imp_line = tot_line - 1
    sub_line = imp_line -1
    xe.apply_line_data(ws,sub_line,subtotal_line_ultimo,True)
    xe.apply_line_data(ws,imp_line,imposto_line,True)
    xe.apply_line_data(ws,tot_line,total_line_data,True)
    wb.save(arquivo_editavel)
    messagebox.showinfo("Sucesso", "Relatório Gerado com sucesso!")
        
    return

def escolhe_arquivo():
    root = tk.Tk()
    root.withdraw()  # Esconde a janela principal do Tkinter
    caminho_arquivo = filedialog.askopenfilename(
        title="Selecione a planilha de orçamentos",
        filetypes=(("Arquivos de texto", "*.xlsx"), ("Todos os arquivos", "*.*"))
    )
    return caminho_arquivo

def criar_aba_categoria(nome_categoria):
    # Frame principal da aba
    frame_aba = ttk.Frame(notebook)
    notebook.add(frame_aba, text=nome_categoria)
    
    # Container com scroll
    container = ttk.Frame(frame_aba)
    container.pack(fill="both", expand=True)
    
    canvas = tk.Canvas(container)
    scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
    scrollable_frame = ttk.Frame(canvas)
    
    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Cabeçalho da tabela
    colunas = ["Item", "Tipo", "Descrição", "Qtde", "Diária", "Subtotal", "Total", "Incluir no Orçamento?"]
    pesos_colunas = [2,2,2,1,1,1,1,1]
    for i, peso in enumerate(pesos_colunas):
        scrollable_frame.columnconfigure(i, weight=peso)

    for i, coluna in enumerate(colunas):
        ttk.Label(scrollable_frame, text=coluna, font=('Arial', 9, 'bold')).grid(
            row=0, column=i, padx=2, pady=5, sticky="ew"
        )
    
    # Frame para os campos de entrada
    frame_campos = ttk.Frame(scrollable_frame)
    frame_campos.grid(row=1, column=0, columnspan=len(colunas), sticky="ew")
    
    # Configurar grid weights
    for i, peso in enumerate(pesos_colunas):
        frame_campos.columnconfigure(i, weight=peso)
    
    # Lista para armazenar as linhas desta categoria
    
    def adicionar_linha():
        global linhas_categoria
        row_index = len(linhas_categoria) + 1
        
        # Frame para uma linha
        frame_linha = ttk.Frame(frame_campos)
        frame_linha.grid(row=row_index, column=0, columnspan=len(colunas), sticky="ew", pady=1)
        
        for i, peso in enumerate(pesos_colunas):
            frame_linha.columnconfigure(i, weight=peso)

        # Campos da linha
        entry_item = ttk.Entry(frame_linha)
        entry_item.grid(row=0, column=0, padx=2, sticky="ew")

        entry_tipo = ttk.Entry(frame_linha)
        entry_tipo.grid(row=0, column=1, padx=2, sticky="ew")

        entry_descricao = ttk.Entry(frame_linha)
        entry_descricao.grid(row=0, column=2, padx=2, sticky="ew")

        entry_qtde = ttk.Entry(frame_linha)
        entry_qtde.grid(row=0, column=3, padx=2, sticky="ew")
        entry_qtde.insert(0, "1")

        entry_diaria = ttk.Entry(frame_linha)
        entry_diaria.grid(row=0, column=4, padx=2, sticky="ew")

        entry_subtotal = ttk.Entry(frame_linha)
        entry_subtotal.grid(row=0, column=5, padx=2, sticky="ew")

        entry_total = ttk.Entry(frame_linha)
        entry_total.grid(row=0, column=6, padx=2, sticky="ew")

        var_incluir = tk.BooleanVar(value=True)
        check_incluir = ttk.Checkbutton(frame_linha, variable=var_incluir)
        check_incluir.grid(row=0, column=7, padx=2)
        
        # Configurar weights para expansão
        for i in range(len(colunas)):
            frame_linha.columnconfigure(i, weight=1)
        
        linha_data = {
            "frame": frame_linha,
            "entries": {
                "categoria": nome_categoria,
                "item": entry_item,
                "tipo": entry_tipo,
                "descricao": entry_descricao,
                "qtde": entry_qtde,
                "diaria": entry_diaria,
                "subtotal": entry_subtotal,
                "total": entry_total,
                "incluir": var_incluir
            }
        }
        
        linhas_categoria.append(linha_data)

    def salvar_categoria():
        # Coletar dados da categoria
        itens_categoria = []
        for linha in linhas_categoria:
            if linha["entries"]["incluir"].get() == 1:
                linha["entries"]["incluir"] = "Sim"
            else:
                linha["entries"]["incluir"] = "Não"
            item_data = {
                "categoria": nome_categoria,
                "item": linha["entries"]["item"].get(),
                "tipo": linha["entries"]["tipo"].get(),
                "descricao": linha["entries"]["descricao"].get(),
                "qtde": linha["entries"]["qtde"].get(),
                "diaria": linha["entries"]["diaria"].get(),
                "subtotal": linha["entries"]["subtotal"].get(),
                "total": linha["entries"]["total"].get(),
                "incluir": linha["entries"]["incluir"]
                }
            itens_categoria.append(item_data)
        dados=[]
        for item in itens_categoria:
            dados.append({
                "Inicio": dados_orcamento['DATA DE INICIO'],
                "Termino": dados_orcamento['DATA DE TÉRMINO'],
                "Evento": dados_orcamento['PROJETO/EVENTO:'],
                "Categoria": item['categoria'],
                "Item": item['item'],
                "tipo": item['tipo'],
                "Quantidade": item['qtde'],
                "Diária": item['diaria'],
                "Subtotal": item['subtotal'],
                "Total": item['total'],
                "Incluir": item['incluir']})
        
        lancamento="LANÇAMENTOS_FORMATADO"
        df_final = pd.DataFrame(dados)
        try:
            df_existente = pd.read_excel(arquivo, sheet_name=lancamento)
            df_final = pd.concat([df_existente,df_final], ignore_index=False)
        except FileNotFoundError:
            pass
        except ValueError:
            pass

        with pd.ExcelWriter(arquivo, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df_final.to_excel(writer, sheet_name=lancamento, index=False)

        messagebox.showinfo("Sucesso", f"Itens da categoria {nome_categoria} salvos com sucesso!")
    
    # Botão para adicionar linha
    btn_adicionar = ttk.Button(scrollable_frame, text="+ Adicionar Linha", command=adicionar_linha)
    btn_adicionar.grid(row=2, column=0, columnspan=2, pady=10, sticky="w")
    
    # Botão para salvar categoria
    btn_salvar_categoria = ttk.Button(scrollable_frame, text=f"Salvar {nome_categoria}", command=salvar_categoria)
    btn_salvar_categoria.grid(row=2, column=2, columnspan=2, pady=10)
    
    # Adicionar primeira linha automaticamente
    adicionar_linha()

def voltar_centro_custo():
    frame_detalhamento.pack_forget()
    frame_centro_custo.pack(fill="both", expand=True)

def finalizar_orcamento():
    messagebox.showinfo("Sucesso", "Orçamento completo salvo com sucesso!")
    limpar_campos()
    resetar_orcamento()

def limpar_campos():
    # Limpar campos do orçamento
    for entry in [entry_cliente, entry_evento, entry_solicitado, entry_telefone,
                  entry_email, entry_gerente, entry_recebido, entry_aprovado,
                  entry_inicio, entry_termino, entry_enviado]:
        entry.delete(0, tk.END)

    # Limpar campos do centro de custo
    for entry in [entry_estrutura, entry_atracoes, entry_alimentos, entry_equipamentos,
                  entry_servicos, entry_equipe, entry_taxas, entry_divulgacao, entry_outros]:
        entry.delete(0, tk.END)

def on_mousewheel(event):
    canvas.yview_scroll(-1 * int(event.delta / 120), "units")

def on_mousewheel_centro_custo(event):
    canvas_centro_custo.yview_scroll(-1 * int(event.delta / 120), "units")

# Janela principal
pasta_atual = Path.cwd()
arquivo = pasta_atual / "ORÇAMENTO DE EVENTOS.xlsx"
if arquivo.exists():
    pass
else:
    arquivo = escolhe_arquivo()
janela = tk.Tk()
janela.title("Sistema de Orçamento")
janela.geometry("800x600")

# ========== FRAME DO ORÇAMENTO ==========
frame_orcamento = ttk.Frame(janela)
frame_orcamento.pack(fill="both", expand=True)

# Container com scroll para o orçamento
frame_container = ttk.Frame(frame_orcamento)
frame_container.pack(fill="both", expand=True)

canvas = tk.Canvas(frame_container)
scrollbar = ttk.Scrollbar(frame_container, orient="vertical", command=canvas.yview)
scrollable_frame = ttk.Frame(canvas)

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind("<MouseWheel>", on_mousewheel)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# Formulário do orçamento no frame rolável
scrollable_frame.columnconfigure(1, weight=1)

# Campos do orçamento
campos_orcamento = [
    ("CLIENTE:", "entry_cliente"),
    ("PROJETO/EVENTO:", "entry_evento"),
    ("SOLICITADO POR:", "entry_solicitado"),
    ("TELEFONE:", "entry_telefone"),
    ("E-MAIL:", "entry_email"),
    ("GERENTE DO PROJETO:", "entry_gerente"),
    ("RECEBIDO EM:", "entry_recebido"),
    ("APROVADO EM:", "entry_aprovado"),
    ("DATA DE INICIO:", "entry_inicio"),
    ("DATA DE TÉRMINO:", "entry_termino"),
    ("Enviado em ", "entry_enviado")
]

entries_orcamento = {}
for i, (label_text, entry_name) in enumerate(campos_orcamento):
    ttk.Label(scrollable_frame, text=label_text).grid(row=i, column=0, padx=5, pady=5, sticky="e")
    entry = ttk.Entry(scrollable_frame, width=40)
    entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
    entries_orcamento[entry_name] = entry

# Atribuir às variáveis globais
entry_cliente = entries_orcamento["entry_cliente"]
entry_evento = entries_orcamento["entry_evento"]
entry_solicitado = entries_orcamento["entry_solicitado"]
entry_telefone = entries_orcamento["entry_telefone"]
entry_email = entries_orcamento["entry_email"]
entry_gerente = entries_orcamento["entry_gerente"]
entry_recebido = entries_orcamento["entry_recebido"]
entry_aprovado = entries_orcamento["entry_aprovado"]
entry_inicio = entries_orcamento["entry_inicio"]
entry_termino = entries_orcamento["entry_termino"]
entry_enviado = entries_orcamento["entry_enviado"]

# Botões do orçamento
button_frame_orcamento = ttk.Frame(scrollable_frame)
button_frame_orcamento.grid(row=len(campos_orcamento), column=0, columnspan=2, pady=20)

ttk.Button(button_frame_orcamento, text="Salvar e Continuar",
          command=salvar_orcamento).pack(side="left", padx=10)
ttk.Button(button_frame_orcamento, text="Limpar",
          command=limpar_campos).pack(side="left", padx=10)

# ========== FRAME DO CENTRO DE CUSTO ==========
frame_centro_custo = ttk.Frame(janela)

# Container com scroll para o centro de custo
frame_container_centro = ttk.Frame(frame_centro_custo)
frame_container_centro.pack(fill="both", expand=True)

canvas_centro_custo = tk.Canvas(frame_container_centro)
scrollbar_centro = ttk.Scrollbar(frame_container_centro, orient="vertical", command=canvas_centro_custo.yview)
scrollable_frame_centro = ttk.Frame(canvas_centro_custo)

scrollable_frame_centro.bind(
    "<Configure>",
    lambda e: canvas_centro_custo.configure(scrollregion=canvas_centro_custo.bbox("all"))
)

canvas_centro_custo.create_window((0, 0), window=scrollable_frame_centro, anchor="nw")
canvas_centro_custo.configure(yscrollcommand=scrollbar_centro.set)
canvas_centro_custo.bind("<MouseWheel>", on_mousewheel_centro_custo)

canvas_centro_custo.pack(side="left", fill="both", expand=True)
scrollbar_centro.pack(side="right", fill="y")

# Formulário do centro de custo
scrollable_frame_centro.columnconfigure(1, weight=1)

# Categorias do centro de custo
categorias = [
    "ESTRUTURA & CENOGRAFIA R$",
    "ATRAÇÕES R$",
    "ALIMENTOS E BEBIDAS R$",
    "LOCAÇÃO DE EQUIPAMENTOS R$",
    "SERVIÇOS R$",
    "EQUIPE/PRODUÇÃO R$",
    "TAXAS/LEGALIZAÇÃO R$",
    "DIVULGAÇÃO R$",
    "OUTROS R$"
]

entries_centro_custo = {}
for i, categoria in enumerate(categorias):
    ttk.Label(scrollable_frame_centro, text=categoria, font=('Arial', 10, 'bold')).grid(row=i, column=0, padx=5, pady=8, sticky="e")
    entry = ttk.Entry(scrollable_frame_centro, width=20)
    entry.grid(row=i, column=1, padx=5, pady=8, sticky="ew")
    entries_centro_custo[categoria] = entry

# Atribuir às variáveis globais
entry_estrutura = entries_centro_custo["ESTRUTURA & CENOGRAFIA R$"]
entry_atracoes = entries_centro_custo["ATRAÇÕES R$"]
entry_alimentos = entries_centro_custo["ALIMENTOS E BEBIDAS R$"]
entry_equipamentos = entries_centro_custo["LOCAÇÃO DE EQUIPAMENTOS R$"]
entry_servicos = entries_centro_custo["SERVIÇOS R$"]
entry_equipe = entries_centro_custo["EQUIPE/PRODUÇÃO R$"]
entry_taxas = entries_centro_custo["TAXAS/LEGALIZAÇÃO R$"]
entry_divulgacao = entries_centro_custo["DIVULGAÇÃO R$"]
entry_outros = entries_centro_custo["OUTROS R$"]

# Botões do centro de custo
button_frame_centro = ttk.Frame(scrollable_frame_centro)
button_frame_centro.grid(row=len(categorias), column=0, columnspan=2, pady=20)

ttk.Button(button_frame_centro, text="Salvar e Continuar",
          command=salvar_centro_custo).pack(side="left", padx=10)
ttk.Button(button_frame_centro, text="Voltar",
          command=voltar_orcamento).pack(side="left", padx=10)

# ========== FRAME DE DETALHAMENTO DE CUSTOS ==========
frame_detalhamento = ttk.Frame(janela)

# Notebook para as abas de categorias
notebook = ttk.Notebook(frame_detalhamento)
notebook.pack(fill="both", expand=True, padx=10, pady=10)

# Botões do frame de detalhamento
button_frame_detalhamento = ttk.Frame(frame_detalhamento)
button_frame_detalhamento.pack(fill="x", pady=10)

ttk.Button(button_frame_detalhamento, text="Novo Orçamento",
          command=finalizar_orcamento).pack(side="left", padx=10)
ttk.Button(button_frame_detalhamento, text="Gerar Relatório",
          command=salvar_relatorio).pack(side="left", padx=10)
ttk.Button(button_frame_detalhamento, text="Voltar",
          command=voltar_centro_custo).pack(side="left", padx=10)

# Iniciar com o frame do orçamento visível
frame_centro_custo.pack_forget()
frame_detalhamento.pack_forget()

janela.mainloop()
